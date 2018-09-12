import os
from openpyxl import load_workbook
from flask import session
from flask_login import current_user, UserMixin
#from app import login


class Quiz(object):
    # get file:
    # with current_app.open_instance_resource('application.cfg') as f:

    def __init__(self, title=None):
        self.title = title
        self._datasets = []
        self.questions = []
        self.submitted_answers = []

    def build_questions(self):
        for dataset in self._datasets:
            for entry in dataset['entries']:
                headers = list(dataset['headers'])
                primary_key = headers.pop(0)
                primary_value = entry.pop(0)
                for index, value in enumerate(entry):
                    self.questions.append({
                        'section': dataset['section'],
                        'known': {primary_key: primary_value},
                        'unknown': headers[index],
                        'correct_answer': value
                    })
                    # self.questions.append({
                    #     'section': dataset['section'],
                    #     'known': {headers[index]: value},
                    #     'unknown': primary_key,
                    #     'correct_answer': primary_value
                    # })

    def load_from_excel(self, filename):
        self.title = os.path.splitext(os.path.basename(filename))[0]
        wb = load_workbook(filename)

        for sheet in wb:
            headers = sheet[1]  # get headers from row 1

            dataset = {
                'section': sheet.title,
                'headers': [x.value for x in headers],
                'entries': []
            }

            for row in sheet.iter_rows(min_row=2):  # iterate from row 2
                dataset['entries'].append([x.value for x in row])

            self._datasets.append(dataset)

        return self


class User(UserMixin):
    def __init__(self):
        self.id = None
        self.username = None
