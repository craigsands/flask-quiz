import errno
import os
import random
import time
from flask import (
    current_app, flash, redirect, render_template, request, url_for)
from flask_login import login_required
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from app import db
from app.models import Question, Quiz, Subject
from app.question import bp
from app.question.forms import UploadQuizForm


@bp.route('/')
@login_required
def index():
    return render_template('question/index.html', title='Questions')


@bp.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    form = UploadQuizForm()

    if request.method == "POST":
        if form.validate_on_submit():
            try:
                f = form.file.data
                filename = secure_filename(f.filename)

                upload_dir = os.path.join(current_app.instance_path, 'uploads')
                try:
                    os.makedirs(upload_dir)
                except OSError as exception:
                    if exception.errno != errno.EEXIST:
                        raise

                filepath = os.path.join(upload_dir, filename)
                f.save(filepath)
                current_app.logger.info('file saved to {}'.format(filepath))

                # collect questions for adding to sample quiz
                questions = []

                # uploading a file creates (num_columns-1)*2*rows questions
                # questions are added to questions table
                # on first upload a quiz is created containing all questions
                # you can also create quizzes by selecting a subject, or by
                # choosing from a list of questions
                wb = load_workbook(filepath)
                for sheet in wb:
                    subject = Subject.query.filter_by(name=sheet.title).first()
                    if not subject:
                        subject = Subject(name=sheet.title)
                        db.session.add(subject)
                        db.session.flush()

                    headers = [x.value for x in sheet[1]]  # get headers from row 1
                    primary_key = headers.pop(0)
                    for row in sheet.iter_rows(min_row=2):
                        values = [x.value for x in row]
                        primary_value = values.pop(0)

                        for index, value in enumerate(values):
                            # Question the value knowing the key
                            question = Question(
                                subject_id=subject.id,
                                prompt='Given the {} is {}, what is the {}?'.format(
                                    primary_key, primary_value, headers[index]
                                ),
                                correct_answer=value
                            )
                            db.session.add(question)
                            questions.append(question)

                            # Question the key, knowing the value
                            inverse = Question(
                                subject_id=subject.id,
                                prompt='Given the {} is {}, what is the {}?'.format(
                                    headers[index], value, primary_key
                                ),
                                correct_answer=primary_value
                            )
                            db.session.add(inverse)
                            questions.append(inverse)

                db.session.commit()
                flash('Congratulations, you just uploaded questions!')

                if form.create_quiz.data:
                    current_millis = int(round(time.time() * 1000))
                    quiz_name = '%s-%s' % (filename, current_millis)
                    quiz = Quiz(name=quiz_name)
                    sample_size = min(len(questions), 10)
                    quiz.questions = random.sample(questions, sample_size)
                    db.session.add(quiz)
                    db.session.commit()
                    flash('Quiz created with %s random questions.' %
                          sample_size)

            finally:
                return redirect(url_for('question.index'))

    return render_template('question/upload.html', form=form)